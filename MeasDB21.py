# coding: UTF-8
import os
import subprocess
import tkinter.ttk
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import Calendar, DateEntry
import numpy as np
import matplotlib.dates as mdates # mdateのインポート
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import japanize_matplotlib 
import matplotlib
import mplcursors
import mysql.connector
import datetime
from dateutil.relativedelta import relativedelta
import openpyxl
import pandas as pd

# Global変数初期値設定
global Mydb, Pptn, Psdate, Pedate, Pexcl, cmb1, ent1, ent2, Fig, Ax
global List_chk, List_devser, List_meas, Pdevl, Pnamel
Mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="den3passwd",
    database="sms")
List_chk = []
Pptn = ''
Pdevl = ['','','','','']
Pnamel = ['','','','','']
Psdate = ""
Pedate = ""
Pexcl = os.getcwd() + "/excel/meas_"
#Win Pexcl = os.getcwd() + "\meas_"
# 
# 計測値リスト作成
def MeasList():
    global Mydb, List_chk, List_devser, List_meas
    # DBから測定値を取得
    mycursor=Mydb.cursor()
    sql="select distinct devser,macaddr,m_dev_place,m_dev_type,sensor,name,m_meas_jname from meas " + \
	    "left join m_device on devser = m_dev_serno left join m_meas on name = m_meas_name " + \
	    "order by devser, sensor;"
    mycursor.execute(sql)
    result = mycursor.fetchall
    List_devser = []
    list_macaddr = []
    list_place = []
    list_devtype = []
    list_sensor = []
    List_meas = []
    list_measname = []
    for i in mycursor:
        List_devser.append(i[0])
        list_macaddr.append(i[1]) 
        list_place.append(i[2])
        list_devtype.append(i[3])
        list_sensor.append(i[4]) 
        List_meas.append(i[5]) 
        list_measname.append(i[6]) 
    num_list = len(List_devser) #リストの数

    #Canvas widgetを生成
    canvas = tkinter.Canvas(frame_1,width=1004,height=120,bg='white') #背景を白に
    canvas.grid(row=1,rowspan=num_list,column=1,columnspan=9)    #x9列分
    #スクロールバー
    vbar=tkinter.ttk.Scrollbar(frame_1,orient=tkinter.VERTICAL) #縦方向
    vbar.grid(row=2,rowspan=num_list,column=9,sticky='ns'+'e')          #9行分の長さで設置
    #スクロールバーの制御をCanvasに通知する処理
    vbar.config(command=canvas.yview)
    #Canvasの可動域をスクロールバーに通知する処理
    canvas.config(yscrollcommand=vbar.set)
    #スクロール可動域＜＝これがないと、どこまでもスクロールされてしまう。
    sc_hgt=int(120/4*(num_list+1))  #スクロールの縦の範囲　リストの数＋ヘッダー分に
    canvas.config(scrollregion=(0,0,1000,sc_hgt))

    #Frameを作成
    frame = tkinter.Frame(canvas,bg='white') #背景を白に
    #frameをcanvasに配置
    canvas.create_window((0,0),window=frame,anchor=tkinter.NW,width=canvas.cget('width'))   #anchor<=NWで左上に寄せる

    #header row=1に設定する文字列 余白は0に
    e0=tkinter.Label(frame,width=5,text='select',background='white')
    e0.grid(row=1,column=0,padx=0,pady=0,ipadx=0,ipady=0) #0列目
    e1=tkinter.Label(frame,width=14,text='DeviceSerialNo',background='white')
    e1.grid(row=1,column=1,padx=0,pady=0,ipadx=0,ipady=0) #1列目
    e2=tkinter.Label(frame,width=19,text='MacAddress',background='white')
    e2.grid(row=1,column=2,padx=0,pady=0,ipadx=0,ipady=0) #2列目
    e3=tkinter.Label(frame,width=12,text='Location',background='white')
    e3.grid(row=1,column=3,padx=0,pady=0,ipadx=0,ipady=0) #4列目
    e4=tkinter.Label(frame,width=16,text='DeviceType',background='white')
    e4.grid(row=1,column=4,padx=0,pady=0,ipadx=0,ipady=0) #5列目
    e5=tkinter.Label(frame,width=12,text='SensorType',background='white')
    e5.grid(row=1,column=5,padx=0,pady=0,ipadx=0,ipady=0) #6列目
    e6=tkinter.Label(frame,width=10,text='Meas',background='white')
    e6.grid(row=1,column=6,padx=0,pady=0,ipadx=0,ipady=0) #6列目
    e7=tkinter.Label(frame,width=16,text='MeasName',background='white')
    e7.grid(row=1,column=7,padx=0,pady=0,ipadx=0,ipady=0) #7列目

    irow = 2
    irow0=2
    erow=num_list+irow0
    while irow < erow:   #リストの数分ループしてLabelとチェックボックスを設置
	    #色の設定
        if irow%2==0:
            color='#cdfff7'  #薄い青
        else:
            color='white'
	
	    #チェックボックスの設置
        bln=tkinter.BooleanVar()
        bln.set(False)           #チェックボックスの初期値
        c = tkinter.Checkbutton(frame,variable = bln,width=5,text='',background='white')
        List_chk.append(bln)
        c.grid(row=irow,column=0,padx=0,pady=0,ipadx=0,ipady=0)  #0列目
	    #Device Serial No
        a1=List_devser[irow-irow0]
        b1=tkinter.Label(frame,width=14,text=a1,background=color)
        b1.grid(row=irow,column=1,padx=0,pady=0,ipadx=0,ipady=0) #1列目
	    #Device MacAddress
        a2=list_macaddr[irow-irow0]
        b2=tkinter.Label(frame,width=19,text=a2,background=color)
        b2.grid(row=irow,column=2,padx=0,pady=0,ipadx=0,ipady=0) #2列目
	    #Device Location
        a3=list_place[irow-irow0]
        b3=tkinter.Label(frame,width=12,text=a3,background=color)
        b3.grid(row=irow,column=3,padx=0,pady=0,ipadx=0,ipady=0) #3列目
	    #Device Type
        a4=list_devtype[irow-irow0]
        b4=tkinter.Label(frame,width=16,text=a4,background=color)
        b4.grid(row=irow,column=4,padx=0,pady=0,ipadx=0,ipady=0) #4列目
	    #Sensor Type
        a5=list_sensor[irow-irow0]
        b5=tkinter.Label(frame,width=12,text=a5,background=color)
        b5.grid(row=irow,column=5,padx=0,pady=0,ipadx=0,ipady=0) #5列目
	    #Meas
        a6=List_meas[irow-irow0]
        b6=tkinter.Label(frame,width=10,text=a6,background=color)
        b6.grid(row=irow,column=6,padx=0,pady=0,ipadx=0,ipady=0) #6列目
	    #Meas Name
        a7=list_measname[irow-irow0]
        b7=tkinter.Label(frame,width=16,text=a7,background=color)
        b7.grid(row=irow,column=7,padx=0,pady=0,ipadx=0,ipady=0) #7列目
	
        irow=irow+1

# 検索条件入力Wiget
def SearchConds():     
    global cmb1, ent1, ent2, Pptn, Psdate, Pedate
    # 集計パターン選択
    lbl0 = Label(frame_2, text='集計パターン：')
    lbl0.grid(row=0, column=0, sticky='wens')
    ptn_list = ['明細', '時', '日', '月','年']
    cmb1 = ttk.Combobox(frame_2, values=ptn_list, justify="center")
    cmb1.set(Pptn)
    cmb1.grid(row=0, column=1, sticky='wens')
    cmb1.bind('<<ComboboxSelected>>',sel_ptn)
    # 表示開始日入力
    lbl1 = Label(frame_2, text='　開始日：')
    lbl1.grid(row=0, column=2, sticky=W)
    ent1 = DateEntry(frame_2)
    ent1.grid(row=0, column=3)
    ent1.set_date(datetime.date.today() + relativedelta(months=-1))
    # 表示終了日入力
    lbl2 = Label(frame_2, text='　終了日：')
    lbl2.grid(row=0, column=4, sticky=W)
    ent2 = DateEntry(frame_2)
    ent2.grid(row=0, column=5)
    ent2.set_date(datetime.date.today())
    # 最新表示指示
    btn1 = Button(frame_2, text='View', command=execg)
    btn1.grid(row=0, column=6, sticky=E)
    # EXCEL DownLoad
    btn2 = Button(frame_2, text='Excel', command=dwnld)
    btn2.grid(row=0, column=7, sticky=E)


# 検索条件取得
def GetConds():     
    global ent1, ent2, Pptn, Psdate, Pedate, Pdevl, Pnamel
    global List_chk, List_devser, List_meas

    Psdate = str(ent1.get_date())   #期間開始日
    Pedate = str(ent2.get_date())   #期間終了日   
    Pdevl = ['','','','','']        #デバイスシリアル№ 
    Pnamel = ['','','','','']       #計測値名称 
    i = 0
    num_list = len(List_devser)
    for ilist in range(num_list):
        if (List_chk[ilist].get()):   #checkbuttonの値
            Pdevl[i] = List_devser[ilist]
            Pnamel[i] = List_meas[ilist]
            i += 1

# 計測値グラフ作成
def MeasPlot():
    global Mydb, Pptn, Pdevl, Pnamel, Psdate, Pedate, Fig, Ax, xdate, yvalue
   # DBから測定値を取得
    mycursor=Mydb.cursor()
    xdate = [[],[],[],[],[]]
    yvalue = [[],[],[],[],[]]
    maxy = 0.0
    miny = 0.0
    ff = True
    sql = ""
    for j in range(5): 
        if Pdevl[j] != '':
            if Pptn == "年":
                sql="select DATE_FORMAT(date,'%Y') , truncate(avg(value),3) from meas " + \
                    "where devser='" + Pdevl[j] + "' and name='" + Pnamel[j] + "' and DATE_FORMAT(date,'%Y-%m-%d/%H')>='" + Psdate + "/00' " + \
                    "and DATE_FORMAT(date,'%Y-%m-%d/%H')<='" + Pedate + "/24' " + \
                    "group by DATE_FORMAT(date,'%Y')"
            elif Pptn == "月":
                sql="select DATE_FORMAT(date,'%Y-%m') , truncate(avg(value),3) from meas " + \
                    "where devser='" + Pdevl[j] + "' and name='" + Pnamel[j] + "' and DATE_FORMAT(date,'%Y-%m-%d/%H')>='" + Psdate + "/00' " + \
                    "and DATE_FORMAT(date,'%Y-%m-%d/%H')<='" + Pedate + "/24' " + \
                    "group by DATE_FORMAT(date,'%Y-%m')"
            elif Pptn == "日":
                sql="select DATE_FORMAT(date,'%Y-%m-%d') , truncate(avg(value),3) from meas " + \
                    "where devser='" + Pdevl[j] + "' and name='" + Pnamel[j] + "' and DATE_FORMAT(date,'%Y-%m-%d/%H')>='" + Psdate + "/00' " + \
                    "and DATE_FORMAT(date,'%Y-%m-%d/%H')<='" + Pedate + "/24' " + \
                    "group by DATE_FORMAT(date,'%Y-%m-%d')"
            elif Pptn == "時":
                sql="select DATE_FORMAT(date,'%Y-%m-%d %H') , truncate(avg(value),3) from meas " + \
                    "where devser='" + Pdevl[j] + "' and name='" + Pnamel[j] + "' and DATE_FORMAT(date,'%Y-%m-%d/%H')>='" + Psdate + "/00' " + \
                    "and DATE_FORMAT(date,'%Y-%m-%d/%H')<='" + Pedate + "/24' " + \
                    "group by DATE_FORMAT(date,'%Y%m%d/%H')"
            else:
                sql="select DATE_FORMAT(date,'%Y-%m-%d %H:%i:%s') , truncate(value,3) from meas " + \
                    "where devser='" + Pdevl[j] + "' and name='" + Pnamel[j] + "' and DATE_FORMAT(date,'%Y-%m-%d/%H')>='" + Psdate + "/00' " + \
                    "and DATE_FORMAT(date,'%Y-%m-%d/%H')<='" + Pedate + "/24' "
            mycursor.execute(sql)
            result = mycursor.fetchall
            for i in mycursor:
                if Pptn=="年":
                    dt = datetime.datetime.strptime(i[0], "%Y")
                elif Pptn=="月":
                    dt = datetime.datetime.strptime(i[0], "%Y-%m")
                elif Pptn=="日":
                    dt = datetime.datetime.strptime(i[0], "%Y-%m-%d")
                elif Pptn=="時":
                    dt = datetime.datetime.strptime(i[0], "%Y-%m-%d %H")
                else:
                    dt = datetime.datetime.strptime(i[0], "%Y-%m-%d %H:%M:%S")
                xdate[j].append(dt)
                yvalue[j].append(i[1])
                if ((i[1] > maxy) or (ff)):
                    maxy = i[1]
                if ((i[1] < miny) or (ff)):
                    miny = i[1]
                ff = False
    if (maxy >= 0):
        maxy = maxy + maxy*0.1
    else :
        mavy = maxy - maxy*0.1
    if (miny >= 0):
        miny = miny - miny*0.1
    else :
        miny = miny + miny*0.1
    # グラフ作成
    Fig = plt.Figure(figsize=[10.24, 3.6])
    Ax = Fig.add_subplot(1,1,1)
    # No select
    if ((Pdevl[0] == "") or (miny == 0 and maxy == 0)):
        Ax.set_title("Data Not Found!!")
    else:
        # y-axis
        Ax.set_ylim(miny,maxy)
        # setting ylabel of graph
        Ax.set_ylabel("Temp(℃) / Humi(%)")
        # Setting count of values in
        plt.tick_params(labelsize=15)
        if Pptn == '年':
            span = pd.to_datetime([Psdate[0:4], Pedate[0:4]])
        elif Pptn == '月':
            span = pd.to_datetime([Psdate[0:7], Pedate[0:7]])
        elif Pptn == '日':
            span = pd.to_datetime([Psdate, Pedate])
        elif Pptn == '時':
            span = pd.to_datetime([Psdate + ' 00', Pedate + ' 23'])
        else:
            span = pd.to_datetime([Psdate, Pedate])
        Ax.set_xlim(span)       
        Fig.autofmt_xdate()
        for j in range(5):
            if Pdevl[j] != '':
                if j == 0:
                    line0 = Ax.plot(xdate[j],yvalue[j],label=Pdevl[j] + ":" + Pnamel[j],marker='o',picker=15)
                    mplcursors.cursor(line0)
                if j == 1:
                    line1 = Ax.plot(xdate[j],yvalue[j],label=Pdevl[j] + ":" + Pnamel[j],marker='o',picker=15)
                    mplcursors.cursor(line1)
                if j == 2:
                    line2 = Ax.plot(xdate[j],yvalue[j],label=Pdevl[j] + ":" + Pnamel[j],marker='o',picker=15)
                    mplcursors.cursor(line2)
                if j == 3:
                    line3 = Ax.plot(xdate[j],yvalue[j],label=Pdevl[j] + ":" + Pnamel[j],marker='o',picker=15)
                    mplcursors.cursor(line3)
                if j == 4:
                    line4 = Ax.plot(xdate[j],yvalue[j],label=Pdevl[j] + ":" + Pnamel[j],marker='o',picker=15)
                    mplcursors.cursor(line4)
                if j == 5:
                    line5 = Ax.plot(xdate[j],yvalue[j],label=Pdevl[j] + ":" + Pnamel[j],marker='o',picker=15)
                    mplcursors.cursor(line5)
        # setting xlabel of graph
        Ax.set_xlabel("Meas Date&Time(yyyymmdd/hh)")
        # setting tile of graph
        Ax.set_title("Environment Information")
        plt.tick_params(labelsize=10)
        plt.grid()
        Ax.legend()

# グラフ表示ボタン
def execg():
    global Fig, ent1, ent2, Psdate, Pedate, Pdevl, Pnamel, List_chk, List_devser, List_meas
    GetConds()
    MeasPlot()
    canvas = FigureCanvasTkAgg(Fig, frame_3)
    canvas.draw()
    canvas.get_tk_widget().grid(row=0, column=0)

# EXCEL作成ボタン
def dwnld():
    global Mydb, Psdate, Pexcl, Pdevl, Pnamel, Pptn

    # 選択条件取得
    GetConds()
    # DBから測定値を取得
    mycursor=Mydb.cursor()
    for j in range(5): 
        if Pdevl[j] != '':
            if Pptn == "年":
                sql="select devser,macaddr,m_dev_place,m_dev_type,sensor,name,m_meas_jname,DATE_FORMAT(date,'%Y') , truncate(avg(value),3) from meas " + \
                    "left join m_device on devser = m_dev_serno left join m_meas on name = m_meas_name " + \
                    "where devser='" + Pdevl[j] + "' and name='" + Pnamel[j] + "' and DATE_FORMAT(date,'%Y-%m-%d/%H')>='" + Psdate + "/00' " + \
                    "and DATE_FORMAT(date,'%Y-%m-%d/%H')<='" + Pedate + "/24' " + \
                    "group by DATE_FORMAT(date,'%Y')"
            elif Pptn == "月":
                sql="select devser,macaddr,m_dev_place,m_dev_type,sensor,name,m_meas_jname, DATE_FORMAT(date,'%Y-%m') , truncate(avg(value),3) from meas " + \
                    "left join m_device on devser = m_dev_serno left join m_meas on name = m_meas_name " + \
                    "where devser='" + Pdevl[j] + "' and name='" + Pnamel[j] + "' and DATE_FORMAT(date,'%Y-%m-%d/%H')>='" + Psdate + "/00' " + \
                    "and DATE_FORMAT(date,'%Y-%m-%d/%H')<='" + Pedate + "/24' " + \
                    "group by DATE_FORMAT(date,'%Y-%m')"
            elif Pptn == "日":
                sql="select devser,macaddr,m_dev_place,m_dev_type,sensor,name,m_meas_jname, DATE_FORMAT(date,'%Y-%m-%d') , truncate(avg(value),3) from meas " + \
                    "left join m_device on devser = m_dev_serno left join m_meas on name = m_meas_name " + \
                    "where devser='" + Pdevl[j] + "' and name='" + Pnamel[j] + "' and DATE_FORMAT(date,'%Y-%m-%d/%H')>='" + Psdate + "/00' " + \
                    "and DATE_FORMAT(date,'%Y-%m-%d/%H')<='" + Pedate + "/24' " + \
                    "group by DATE_FORMAT(date,'%Y-%m-%d')"
            elif Pptn == "時":
                sql="select devser,macaddr,m_dev_place,m_dev_type,sensor,name,m_meas_jname, DATE_FORMAT(date,'%Y-%m-%d %H') , truncate(avg(value),3) from meas " + \
                    "left join m_device on devser = m_dev_serno left join m_meas on name = m_meas_name " + \
                    "where devser='" + Pdevl[j] + "' and name='" + Pnamel[j] + "' and DATE_FORMAT(date,'%Y-%m-%d/%H')>='" + Psdate + "/00' " + \
                    "and DATE_FORMAT(date,'%Y-%m-%d/%H')<='" + Pedate + "/24' " + \
                    "group by DATE_FORMAT(date,'%Y%m%d/%H')"
            else:
                sql="select devser,macaddr,m_dev_place,m_dev_type,sensor,name,m_meas_jname, DATE_FORMAT(date,'%Y-%m-%d %H:%i:%s') , truncate(value,3) from meas " + \
                    "left join m_device on devser = m_dev_serno left join m_meas on name = m_meas_name " + \
                    "where devser='" + Pdevl[j] + "' and name='" + Pnamel[j] + "' and DATE_FORMAT(date,'%Y-%m-%d/%H')>='" + Psdate + "/00' " + \
                    "and DATE_FORMAT(date,'%Y-%m-%d/%H')<='" + Pedate + "/24' "
            mycursor.execute(sql)
            #result = mycursor.fetchall
            # EXCEL作成
            colnames = mycursor.column_names
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            i = 0
            for n in colnames:
                i += 1    
                sheet.cell(row=1, column=i).value = n
            rows = mycursor.fetchall()
            i = 1
            for rowx in rows:
                i += 1
                k = 0
                for r in rowx:
                    k += 1
                    sheet.cell(row=i, column=k).value = r
            excel = Pexcl + Pnamel[j] + datetime.datetime.now().strftime('%y%m%d%H%M%S') + '.xlsx'
            workbook.save(excel)    
    # excelを開く
    #Win subprocess.Popen(['start', excel], shell=True)
    subprocess.Popen(['see', excel])
    subprocess.wait()

# 30秒毎にグラフ更新
def timer():
    execg()
    root.after(30000,timer)

# When windows is closed.
def _destroyWindow():
    root.quit()
    root.destroy()

# 集計パターン選択
def sel_ptn(event):
    global Pptn, cmb1
    Pptn = cmb1.get()
# 
# メイン
# Windowの設定
root = Tk()
root.title("計測値　グラフ")
root.geometry()
# Frameの設定
frame_1 = Frame(root, bd=4, relief=GROOVE)
frame_2 = Frame(root, bd=4, relief=GROOVE)
frame_3 = Frame(root, bd=4, relief=GROOVE)
# widgetの配置
frame_1.grid(row=0, column=0)
frame_2.grid(row=1, column=0)
frame_3.grid(row=2, column=0)

# 計測値リスト表示
MeasList()

# 検索条件入力widget表示
SearchConds()

# 画面初期表示
execg()

# タイマー起動
#root.after(10000,timer)
# ＬＯＯＰ
root.mainloop()

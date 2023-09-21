import os
import subprocess
from tkinter import *
from tkinter import ttk
import numpy as np
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg 
import mysql.connector
import datetime
import openpyxl

# Global変数初期値設定
global Pdev, Pname, Pdate, Pexcl
Pdev = '30.83.98.A4.9F.E8'
Pname = '*all'
Pdate = '20230101'
Pexcl = os.getcwd() + "/excel/meas_"
#Win Pexcl = os.getcwd() + "\meas_"     


# 計測値グラフ作成
def MeasPlot():
    global Pdev, Pname, Pdate
    if Pdev == '':
        Pdev='30.83.98.A4.9F.E8'
    if Pname == '':
        Pname='*all'
    if Pdate == '':
        Pdate=datetime.datetime.now().strftime('%Y%m%d')
    # DBから測定値を取得
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="den3passwd",
        database="sms_o")
    mycursor=mydb.cursor()
    if Pname == '*all' or Pname == 'temp':
        sql="select DATE_FORMAT(date,'%Y%m%d/%H') , avg(value) from meas " + \
            "where devno='" + Pdev + "' and name='temp' and DATE_FORMAT(date,'%Y%m%d/%H')>='" + Pdate + "/00' " + \
            "group by DATE_FORMAT(date,'%Y%m%d/%H')"
        mycursor.execute(sql)
        result = mycursor.fetchall
        date1 = []
        value1 = []
        for i in mycursor:
            date1.append(i[0])
            value1.append(i[1]) 
    if Pname == '*all' or Pname == 'humi':
        sql="select DATE_FORMAT(date,'%Y%m%d/%H'), avg(value) from meas " + \
            "where devno='" + Pdev + "' and name='humi' and DATE_FORMAT(date,'%Y%m%d/%H')>='" + Pdate + "/00' " + \
            "group by DATE_FORMAT(date,'%Y%m%d/%H')"
        mycursor.execute(sql)
        result = mycursor.fetchall
        date2 = []
        value2 = []
        for i in mycursor:
            date2.append(i[0])
            value2.append(i[1])
    if Pname == '*all' or Pname == 'appt':
        sql="select DATE_FORMAT(date,'%Y%m%d/%H'), avg(value) from meas " + \
            "where devno='" + Pdev + "' and name='appt' and DATE_FORMAT(date,'%Y%m%d/%H')>='" + Pdate + "/00' " + \
            "group by DATE_FORMAT(date,'%Y%m%d/%H')"
        mycursor.execute(sql)
        result = mycursor.fetchall
        date3 = []
        value3 = []
        for i in mycursor:
            date3.append(i[0])
            value3.append(i[1])
    # グラフ作成
    fig = plt.Figure()
    ax = fig.add_subplot(1,1,1)
    fig.autofmt_xdate()
    if Pname == '*all' or Pname == 'temp':
        ax.plot(date1,value1,label='Temp')
    if Pname == '*all' or Pname == 'humi':
        ax.plot(date2,value2,label='Humi')
    if Pname == '*all' or Pname == 'appt':
        ax.plot(date3,value3,label='appt')
    # Setting count of values in
    # y-axis
    ax.set_ylim(0,100)
    # setting xlabel of graph
    ax.set_xlabel("Meas Date&Time(yyyymmdd/hh)")
    # setting ylabel of graph
    ax.set_ylabel("Temp(℃) / Humi(%)")
    # setting tile of graph
    ax.set_title("Environment Information(" + Pdev +")")
    ax.legend()
    return fig

# デバイス選択
def sel_dev(event):
    global Pdev
    Pdev = cmb1.get()

# 測定値選択
def sel_name(event):
    global Pname
    Pname =  cmb2.get()

# グラフ表示ボタン
def exec():
    global Pdate
    Pdate = ent1.get()
    fig = MeasPlot()
    canvas = FigureCanvasTkAgg(fig, frame_2)
    canvas.draw()
    canvas.get_tk_widget().grid(row=0, column=0)

# EXCEL作成ボタン
def dwnld():
    global Pdev, Pname, Pdate, Pexcl
    mname=''
    if Pname != '' and Pname != '*all':
        mname = " and name = '" + Pname + "' "
    # DBから測定値を取得
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="den3passwd",
        database="sms")
    mycursor=mydb.cursor()
    sql="select userId, devtype, devno, sensor, DATE_FORMAT(date,'%Y/%m/%d') as DATE, \
        DATE_FORMAT(date,'%H') as TIME, name,  avg(value) \
        from meas " + \
        "where devno = '" + Pdev + "' " + mname + \
        "and DATE_FORMAT(date,'%Y%m%d/%H') >= '" + Pdate + "/00' " + \
        "group by userId, devtype, devno, sensor, DATE_FORMAT(date,'%Y/%m/%d'), DATE_FORMAT(date,'%H'), name"
    mycursor.execute(sql)
    # EXCEL作成
    colnames = mycursor.column_names
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    i = 0
    for n in colnames:
        i += 1    
        sheet.cell(row=1, column=i).value = n
    rows = mycursor.fetchall()
    i = 1
    for rowx in rows:
        i += 1
        j = 0
        for r in rowx:
            j += 1
            sheet.cell(row=i, column=j).value = r
    excel = Pexcl + datetime.datetime.now().strftime('%y%m%d%H%M%S') + '.xlsx'
    workbook.save(excel)
    # excelを開く
    #Win subprocess.Popen(['start', excel], shell=True)
    subprocess.Popen(['see', excel])
    subprocess.wait()

# 10秒毎にグラフ更新
def timer():
    exec()
    root.after(10000,timer)

# When windows is closed.
def _destroyWindow():
    root.quit()
    root.destroy()

# Windowの設定
root = Tk()
root.title("計測値　グラフ")
root.geometry()

# Frameの設定
frame_1 = Frame(root, bd=4, relief=GROOVE)
frame_2 = Frame(root, bd=4, relief=GROOVE)

# widgetの設定
# デバイスID選択
lbl1 = Label(frame_1, text='デバイスID：')
lbl1.grid(row=0, column=0, sticky='wens')
dev_list = ['30.83.98.A4.9F.E8', '9C.9C.1F.D1.62.08']
cmb1 = ttk.Combobox(frame_1, values=dev_list, justify="center")
cmb1.set(Pdev)
cmb1.grid(row=0, column=1, sticky='wens')
cmb1.bind('<<ComboboxSelected>>',sel_dev)
# 測定値選択
lbl2 = Label(frame_1, text='　測定値：')
lbl2.grid(row=0, column=2, sticky='wens')
name_list = ['*all', 'temp', 'humi', 'appt']
cmb2 = ttk.Combobox(frame_1, values=name_list, justify="center", state='readonly')
cmb2.set(Pname)
cmb2.grid(row=0, column=3, sticky='wens')
cmb2.bind('<<ComboboxSelected>>',sel_name)
# 表示開始日入力
lbl3 = Label(frame_1, text='　開始日：')
lbl3.grid(row=0, column=4)
ent1 = Entry(frame_1)
ent1.insert(END, Pdate)
ent1.grid(row=0, column=5)
# 最新表示指示
btn1 = Button(frame_1, text='View', command=exec)
#btn1.grid(row=1, column=0, columnspan=6, sticky=' padx=10, pady=10)
btn1.grid(row=1, column=0, columnspan=4, sticky='wens')
# EXCEL DownLoad
btn2 = Button(frame_1, text='Excel', command=dwnld)
btn2.grid(row=1, column=4, columnspan=3, sticky='wens')

# 表示
fig = MeasPlot()
canvas = FigureCanvasTkAgg(fig, frame_2)
# widgetの配置
frame_1.grid(row=0, column=0, sticky=W + E)
frame_2.grid(row=1, column=0)
canvas.get_tk_widget().grid(row=0, column=0)
# タイマー起動
root.after(10000,timer)
# ＬＯＯＰ
root.mainloop()
